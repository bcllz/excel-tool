"""
Excel 百宝箱 — 智能自动化处理工具
=====================================
基于 Streamlit 的 Excel 处理 Web 应用，共 10 个功能，分 3 大分类。
所有数据在内存中处理，不上传/保存任何用户文件到服务器。
"""

from __future__ import annotations

import io
import zipfile
from datetime import datetime, time
from typing import Optional

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# ===========================================================================
# 页面配置
# ===========================================================================
st.set_page_config(
    page_title="Excel 百宝箱",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ===========================================================================
# CSS
# ===========================================================================
st.markdown("""
<style>
    .main-header { text-align: center; padding: 1rem 0 0.5rem 0; }
    .main-header h1 { font-size: 2.4rem; font-weight: 700; color: #1f77b4; }
    .main-header p { color: #666; font-size: 1rem; }
    .stButton > button { border-radius: 8px; font-weight: 600; }
    .func-active { border: 2px solid #1f77b4 !important; background-color: #e8f4fd !important; }
    .step-label { color: #1f77b4; font-weight: 700; font-size: 1.1rem; margin-bottom: 0.3rem; }
</style>
""", unsafe_allow_html=True)

# ===========================================================================
# Session State 初始化
# ===========================================================================
if "current_function" not in st.session_state:
    st.session_state.current_function = None
if "result_bytes" not in st.session_state:
    st.session_state.result_bytes = None
if "result_ext" not in st.session_state:
    st.session_state.result_ext = ".xlsx"
if "result_filename" not in st.session_state:
    st.session_state.result_filename = "result.xlsx"
if "success_msg" not in st.session_state:
    st.session_state.success_msg = ""

# ===========================================================================
# ===========================================================================
# 工具函数
# ===========================================================================
# ===========================================================================

# ---------------------------------------------------------------------------
# 已有功能：去重
# ---------------------------------------------------------------------------
def do_dedup(df: pd.DataFrame, columns: list[str]) -> tuple[pd.DataFrame, int]:
    before = len(df)
    df = df.drop_duplicates(subset=columns, keep="first")
    return df, before - len(df)

# ---------------------------------------------------------------------------
# 已有功能：拆分
# ---------------------------------------------------------------------------
def do_split(df: pd.DataFrame, column: str, mode: str) -> io.BytesIO:
    groups = df.groupby(column)
    zip_buf = io.BytesIO()

    if mode == "单文件多Sheet":
        inner = io.BytesIO()
        with pd.ExcelWriter(inner, engine="openpyxl") as w:
            for val, grp in groups:
                grp.to_excel(w, sheet_name=str(val)[:31], index=False)
        inner.seek(0)
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"拆分_{column}.xlsx", inner.read())
    else:
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for val, grp in groups:
                b = io.BytesIO()
                grp.to_excel(b, index=False, engine="openpyxl")
                b.seek(0)
                safe = str(val).replace("/", "_").replace("\\", "_")[:50]
                zf.writestr(f"{safe}.xlsx", b.read())
    zip_buf.seek(0)
    return zip_buf

# ---------------------------------------------------------------------------
# 已有功能：合并
# ---------------------------------------------------------------------------
def do_merge(files) -> Optional[pd.DataFrame]:
    dfs = []
    for f in files:
        try:
            xl = pd.ExcelFile(f)
            for sn in xl.sheet_names:
                d = pd.read_excel(xl, sheet_name=sn)
                if not d.empty:
                    dfs.append(d)
        except Exception:
            continue
    return pd.concat(dfs, ignore_index=True) if dfs else None

# ---------------------------------------------------------------------------
# 已有功能：公式列
# ---------------------------------------------------------------------------
def do_formula(df: pd.DataFrame, col_name: str, formula: str) -> io.BytesIO:
    df = df.copy()
    df[col_name] = "公式在 Excel 中计算"

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    for ci, cn in enumerate(df.columns, 1):
        ws.cell(row=1, column=ci, value=cn)
    fidx = list(df.columns).index(col_name) + 1

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, cn in enumerate(df.columns, 1):
            ws.cell(row=ri, column=ci, value=formula if ci == fidx else row[cn])

    for ci in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(
            max(len(str(ws.cell(1, ci).value or "")), 12) + 4, 40
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------------------------------------------------------------------------
# 新功能 1：数据筛选
# ---------------------------------------------------------------------------
def do_filter(df: pd.DataFrame, column: str, op: str, value: str) -> pd.DataFrame:
    """按条件筛选行。"""
    if column not in df.columns:
        raise ValueError(f"列 '{column}' 不存在")

    # 尝试转换数值
    try:
        v = float(value)
    except ValueError:
        v = value

    if op == "等于 (=)":
        return df[df[column] == v]
    elif op == "不等于 (!=)":
        return df[df[column] != v]
    elif op == "大于 (>)":
        return df[df[column].astype(float) > float(value)]
    elif op == "小于 (<)":
        return df[df[column].astype(float) < float(value)]
    elif op == "大于等于 (>=)":
        return df[df[column].astype(float) >= float(value)]
    elif op == "小于等于 (<=)":
        return df[df[column].astype(float) <= float(value)]
    elif op == "包含":
        return df[df[column].astype(str).str.contains(str(value), na=False)]
    elif op == "不包含":
        return df[~df[column].astype(str).str.contains(str(value), na=False)]
    elif op == "为空":
        return df[df[column].isna()]
    elif op == "不为空":
        return df[df[column].notna()]
    else:
        return df

# ---------------------------------------------------------------------------
# 新功能 2：条件格式
# ---------------------------------------------------------------------------
def do_conditional_format(
    df: pd.DataFrame, column: str,
    cond: str, threshold: str, color: str
) -> io.BytesIO:
    """对符合条件的单元格着色，输出带格式的 Excel。"""
    color_map = {
        "红色": "FF6B6B", "绿色": "51CF66", "黄色": "FFD43B",
        "蓝色": "74C0FC", "橙色": "FF922B", "紫色": "B197FC",
    }
    hex_color = color_map.get(color, "FF6B6B")

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    # 写表头
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for ci, cn in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    # 写数据
    highlight_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    col_idx = list(df.columns).index(column) + 1

    try:
        thr_val = float(threshold)
    except ValueError:
        thr_val = threshold

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, cn in enumerate(df.columns, 1):
            cell = ws.cell(row=ri, column=ci, value=row[cn])
            cell.alignment = Alignment(horizontal="center")

        # 判断是否高亮
        cell_val = row[column]
        highlight = False
        try:
            if cond == "大于":
                highlight = float(cell_val) > thr_val
            elif cond == "小于":
                highlight = float(cell_val) < thr_val
            elif cond == "等于":
                highlight = str(cell_val) == str(threshold)
            elif cond == "大于等于":
                highlight = float(cell_val) >= thr_val
            elif cond == "小于等于":
                highlight = float(cell_val) <= thr_val
            elif cond == "包含":
                highlight = str(threshold).lower() in str(cell_val).lower()
        except (ValueError, TypeError):
            pass

        if highlight:
            ws.cell(row=ri, column=col_idx).fill = highlight_fill

    for ci in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------------------------------------------------------------------------
# 新功能 3：数据匹配对比
# ---------------------------------------------------------------------------
def do_match(
    df1: pd.DataFrame, df2: pd.DataFrame,
    col1: str, col2: str, how: str
) -> pd.DataFrame:
    """两张表按指定列匹配（类似 VLOOKUP）。"""
    # 给右表的列加后缀避免重名
    result = df1.merge(
        df2, left_on=col1, right_on=col2,
        how=how, suffixes=("_左表", "_右表")
    )
    return result

# ---------------------------------------------------------------------------
# 新功能 4：透视汇总
# ---------------------------------------------------------------------------
def do_pivot(
    df: pd.DataFrame, index_col: str,
    value_col: str, agg_func: str
) -> pd.DataFrame:
    """生成透视表。"""
    agg_map = {"求和": "sum", "计数": "count", "平均值": "mean",
               "最大值": "max", "最小值": "min"}
    pivoted = pd.pivot_table(
        df, index=index_col, values=value_col,
        aggfunc=agg_map[agg_func]
    ).reset_index()
    pivoted.columns = [str(c) for c in pivoted.columns]
    return pivoted

# ---------------------------------------------------------------------------
# 新功能 5：自动汇总计算（带小计和合计）
# ---------------------------------------------------------------------------
def do_auto_summary(
    df: pd.DataFrame, group_col: str, value_cols: list[str]
) -> io.BytesIO:
    """按某列分组，每组插入小计行，末尾插入合计行。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总报告"

    header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    grandtotal_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    all_cols = list(df.columns)
    # 确保 value_cols 是数值
    for vc in value_cols:
        df[vc] = pd.to_numeric(df[vc], errors="coerce").fillna(0)

    # 写表头
    for ci, cn in enumerate(all_cols, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    row_ptr = 2
    groups = df.groupby(group_col, sort=False)
    totals = {vc: 0.0 for vc in value_cols}

    for grp_name, grp_df in groups:
        # 写数据行
        for _, data_row in grp_df.iterrows():
            for ci, cn in enumerate(all_cols, 1):
                c = ws.cell(row=row_ptr, column=ci, value=data_row[cn])
                c.font = normal_font
                c.alignment = Alignment(horizontal="center")
                c.border = thin_border
            row_ptr += 1

        # 写小计行
        for ci, cn in enumerate(all_cols, 1):
            c = ws.cell(row=row_ptr, column=ci)
            c.fill = subtotal_fill
            c.font = bold_font
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
            if ci == 1:
                c.value = f"📌 {grp_name} 小计"
            elif cn in value_cols:
                sub_val = round(grp_df[cn].sum(), 2)
                totals[cn] += sub_val
                c.value = sub_val
        row_ptr += 1

    # 写合计行
    for ci, cn in enumerate(all_cols, 1):
        c = ws.cell(row=row_ptr, column=ci)
        c.fill = grandtotal_fill
        c.font = Font(bold=True, size=12)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        if ci == 1:
            c.value = "💰 总计 / Grand Total"
        elif cn in value_cols:
            c.value = round(totals[cn], 2)

    for ci in range(1, len(all_cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------------------------------------------------------------------------
# 新功能 6：考勤工资计算
# ---------------------------------------------------------------------------
def do_attendance_salary(
    df: pd.DataFrame,
    work_start: str,
    work_end: str,
    late_penalty: float,
    absent_penalty: float,
    default_salary: float,
    # ---- 列映射 ----
    name_col: str = "姓名",
    date_col: str = "日期",
    start_col: str = "上班打卡",
    end_col: str = "下班打卡",
    salary_col: Optional[str] = None,
) -> io.BytesIO:
    """
    计算考勤并核算工资。
    通过列映射参数适配不同公司的考勤表格式：
      name_col   → 员工姓名列
      date_col   → 日期列
      start_col  → 上班打卡时间列
      end_col    → 下班打卡时间列
      salary_col → 基本工资列（可选，没有则用 default_salary）
    """
    work_start_t = datetime.strptime(work_start, "%H:%M").time()
    work_end_t = datetime.strptime(work_end, "%H:%M").time()

    # ---- 解析打卡时间 ----
    def parse_t(v):
        if pd.isna(v):
            return None
        if isinstance(v, time):
            return v
        if isinstance(v, datetime):
            return v.time()
        s = str(v).strip()
        # 支持各种时间格式
        for fmt in ["%H:%M", "%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                    "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%H:%M:%S.%f"]:
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                continue
        return None

    df = df.copy()

    # 用映射后的列名来操作
    df["__start_t__"] = df[start_col].apply(parse_t)
    df["__end_t__"] = df[end_col].apply(parse_t)

    # ---- 按姓名汇总 ----
    employees = df[name_col].unique()
    results = []
    for emp in employees:
        edf = df[df[name_col] == emp]

        # 日期去重计数
        working_days = edf[date_col].nunique()
        late_count = 0
        early_count = 0
        absent_count = 0

        for _, row in edf.iterrows():
            st_ = row["__start_t__"]
            et_ = row["__end_t__"]

            # 全天没打卡 = 缺勤
            if st_ is None and et_ is None:
                absent_count += 1
                continue

            if st_ and st_ > work_start_t:
                late_count += 1
            if et_ and et_ < work_end_t:
                early_count += 1

        # 基本工资：优先用表中的列，否则用默认值
        if salary_col and salary_col in edf.columns:
            base = edf[salary_col].iloc[0]
        else:
            base = default_salary
        try:
            base = float(base)
        except (ValueError, TypeError):
            base = default_salary

        late_deduct = round(late_count * late_penalty, 2)
        absent_deduct = round(absent_count * absent_penalty, 2)
        net_salary = round(base - late_deduct - absent_deduct, 2)

        results.append({
            "姓名": emp,
            "出勤天数": working_days,
            "迟到次数": late_count,
            "早退次数": early_count,
            "缺勤天数": absent_count,
            "基本工资": base,
            "迟到扣款": late_deduct,
            "缺勤扣款": absent_deduct,
            "实发工资": net_salary,
        })

    result_df = pd.DataFrame(results)

    # ---- 写入 Excel ----
    wb = Workbook()
    ws = wb.active
    ws.title = "工资核算"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    money_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    warning_font = Font(bold=True, color="CC0000", size=11)

    for ci, cn in enumerate(result_df.columns, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for ri, (_, row) in enumerate(result_df.iterrows(), 2):
        for ci, cn in enumerate(result_df.columns, 1):
            c = ws.cell(row=ri, column=ci, value=row[cn])
            c.alignment = Alignment(horizontal="center")
            if cn in ("基本工资", "迟到扣款", "缺勤扣款", "实发工资"):
                c.number_format = '¥#,##0.00'
                c.fill = money_fill
            # 扣款列如果 > 0 标红
            if cn in ("迟到扣款", "缺勤扣款") and row[cn] > 0:
                c.font = warning_font

    for ci in range(1, len(result_df.columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------------------------------------------------------------------------
# 通用：DataFrame → Excel 字节流
# ---------------------------------------------------------------------------
def df_to_excel(df: pd.DataFrame) -> io.BytesIO:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="结果")
    out.seek(0)
    return out

# ---------------------------------------------------------------------------
# 通用：考勤模板生成
# ---------------------------------------------------------------------------
def make_attendance_template() -> io.BytesIO:
    """生成考勤模板（包含标准列名 + 一张说明Sheet）。"""
    wb = Workbook()
    # ---- Sheet 1: 填数据 ----
    ws = wb.active
    ws.title = "考勤数据"

    headers = ["姓名", "日期", "上班打卡", "下班打卡", "基本工资", "备注"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center")

    # 示例数据
    sample = [
        ["张三","2026-08-01","08:50","18:05",8000,"正常"],
        ["张三","2026-08-02","09:15","18:00",8000,"迟到"],
        ["李四","2026-08-01","08:45","17:50",6500,"早退"],
        ["李四","2026-08-02","","",6500,"缺勤"],
        ["王五","2026-08-01","08:30","18:10",10000,"全勤"],
    ]
    for ri, row in enumerate(sample, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="center")

    for i, w in enumerate([10, 14, 12, 12, 12, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 2: 使用说明 ----
    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions["A"].width = 50
    guide = [
        ["📖 考勤模板使用说明", ""],
        ["", ""],
        ["1. 列名可以不同", "如果你的表列名叫「员工姓名」「签到时间」等，上传后在工具里映射即可"],
        ["2. 日期格式", "2026-08-01 或 2026/08/01 都可以"],
        ["3. 打卡时间格式", "08:55 或 08:55:30 或 2026-08-01 08:55:00 都支持"],
        ["4. 基本工资列", "如果没有，工具会使用你设置的默认工资"],
        ["5. 缺勤判断", "当天上班+下班打卡都为空，算缺勤 1 天"],
        ["6. 支持的公司考勤导出格式举例", ""],
        ["  钉钉导出", "姓名 → 姓名列, 考勤日期 → 日期, 签到时间 → 上班打卡, 签退时间 → 下班打卡"],
        ["  企业微信", "姓名 → 姓名列, 日期 → 日期, 上班 → 上班打卡, 下班 → 下班打卡"],
        ["  飞书", "员工 → 姓名列, 日期 → 日期, 签到 → 上班打卡, 签退 → 下班打卡"],
        ["  自定义系统", "只要有姓名/日期/上下班时间这四类列，映射一下就能用"],
    ]
    for ri, (a, b) in enumerate(guide, 1):
        ws2.cell(row=ri, column=1, value=a).font = Font(bold=True, size=11)
        ws2.cell(row=ri, column=2, value=b).font = Font(size=11)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ===========================================================================
# ===========================================================================
# 页面 UI
# ===========================================================================
# ===========================================================================

# ---------- 标题 ----------
st.markdown(
    """<div class="main-header">
        <h1>📊 Excel 百宝箱</h1>
        <p>上传 → 选功能 → 一键处理 → 下载 · 隐私安全 · 纯内存运算 · 10 大功能</p>
    </div>""",
    unsafe_allow_html=True,
)
st.divider()

# ===========================================================================
# 功能分类选择
# ===========================================================================
st.markdown('<p class="step-label">🧰 第一步：选择功能分类</p>', unsafe_allow_html=True)

CATEGORIES = {
    "📊 数据处理": {
        "dedup":  ("🔍 一键去重", "按指定列删除重复行"),
        "split":  ("✂️ 按列拆分", "按某列拆成多个 Sheet/文件"),
        "merge":  ("🔗 数据合并", "多文件纵向拼接为总表"),
        "filter": ("🎯 数据筛选", "按条件过滤行（大于/包含/为空等）"),
        "match":  ("🔗 匹配对比", "两张表按列关联（类 VLOOKUP）"),
    },
    "📈 分析汇总": {
        "pivot":  ("📐 透视汇总", "按类别求和/平均/计数"),
        "summary":("📋 自动汇总", "分组小计 + 合计，格式化输出"),
        "condfmt":("🎨 条件格式", "符合条件的单元格着色标注"),
    },
    "📋 办公模板": {
        "formula":("🧮 公式列", "新增公式列，Excel 打开自动计算"),
        "attend": ("👔 考勤工资", "打卡数据算迟到/缺勤/实发工资"),
    },
}

# 扁平化所有功能
ALL_FUNCTIONS = {}
for cat_name, funcs in CATEGORIES.items():
    for func_id, (label, desc) in funcs.items():
        ALL_FUNCTIONS[func_id] = {"label": label, "desc": desc, "category": cat_name}

cat_keys = list(CATEGORIES.keys())
# 默认选中当前功能所在的分类
default_cat_idx = 0
if st.session_state.current_function:
    cat_of_func = ALL_FUNCTIONS[st.session_state.current_function]["category"]
    if cat_of_func in cat_keys:
        default_cat_idx = cat_keys.index(cat_of_func)

category = st.radio(
    "选择分类", cat_keys,
    horizontal=True, index=default_cat_idx,
    label_visibility="collapsed",
)

# ---------- 功能按钮 ----------
st.markdown('<p class="step-label">🔧 第二步：点击功能按钮</p>', unsafe_allow_html=True)

funcs_in_cat = CATEGORIES[category]
func_ids = list(funcs_in_cat.keys())
cols = st.columns(len(func_ids))

clicked_func = None
for idx, (fid, (label, desc)) in enumerate(funcs_in_cat.items()):
    with cols[idx]:
        btn_label = label
        if st.button(btn_label, key=f"btn_{fid}", use_container_width=True,
                     help=desc,
                     type="primary" if st.session_state.current_function == fid else "secondary"):
            clicked_func = fid

if clicked_func:
    st.session_state.current_function = clicked_func
    st.session_state.result_bytes = None
    st.session_state.success_msg = ""
    st.rerun()

cur_func = st.session_state.current_function
func_info = ALL_FUNCTIONS.get(cur_func, {})
if cur_func:
    st.success(f"📌 当前功能：{func_info.get('label', cur_func)} — {func_info.get('desc', '')}")

st.divider()

# ===========================================================================
# 文件上传区
# ===========================================================================
st.markdown('<p class="step-label">📁 第三步：上传文件</p>', unsafe_allow_html=True)

df = None
df2 = None
uploaded_files = []

if cur_func == "merge":
    uf = st.file_uploader("上传多个 Excel 文件（可多选）", type=["xlsx", "xls"],
                          accept_multiple_files=True, key="upload")
    uploaded_files = uf or []
elif cur_func == "match":
    c1, c2 = st.columns(2)
    with c1:
        f1 = st.file_uploader("上传主表（左表）", type=["xlsx", "xls"], key="match_left")
    with c2:
        f2 = st.file_uploader("上传对照表（右表）", type=["xlsx", "xls"], key="match_right")
    if f1:
        df = pd.read_excel(f1)
    if f2:
        df2 = pd.read_excel(f2)
elif cur_func:
    if cur_func == "attend":
        # 考勤功能额外提供模板下载
        tmpl = make_attendance_template()
        st.download_button(
            "📥 下载考勤模板（按此格式填写后上传）",
            data=tmpl, file_name="考勤模板.xlsx",
            mime="application/octet-stream",
        )
    uf = st.file_uploader("上传 Excel 文件", type=["xlsx", "xls"],
                          accept_multiple_files=False, key="upload_single")
    if uf:
        df = pd.read_excel(uf)

# ===========================================================================
# 使用指南
# ===========================================================================
GUIDES = {
    "dedup": """
### 🔍 一键去重 — 怎么用？

**干什么：** 表格里有重复的行？一键删掉，只留第一条。

**三步搞定：**
1. 📁 上传你的 Excel 文件
2. ✅ 勾选要用来判断重复的列（比如选了「姓名」+「日期」，那姓名和日期都相同的行就算重复）
3. 🚀 点执行 → 下载

**💡 举个栗子：**
```
张三 | 北京 | 100
李四 | 上海 | 200
张三 | 北京 | 100  ← 跟第一条一模一样，会被删掉
```
选「姓名」「城市」「金额」三列 → 去重后只剩 2 行。

**⚠️ 注意：** 只保留第一次出现的行，后面的重复行直接删除。
""",
    "split": """
### ✂️ 按列拆分 — 怎么用？

**干什么：** 一张大表按某一列拆成多个小表。比如全国销售数据按「省份」拆——每个省一份。

**三步搞定：**
1. 📁 上传 Excel 文件
2. 🔽 选择拆分依据列（比如「省份」→ 广东/浙江/江苏各一份）
3. 🔘 选模式：单文件多 Sheet（一个 Excel 多个标签页）还是多个文件打包 ZIP
4. 🚀 点执行 → 下载

**💡 举个栗子：**
上传一个包含 3 个省份、8 行数据的表 → 选「省份」列 → 拆成广东.xlsx、浙江.xlsx、江苏.xlsx → 打包成 ZIP 下载。

**⚠️ 注意：** 拆分依据列有多少个不同值，就拆成多少份。
""",
    "merge": """
### 🔗 数据合并 — 怎么用？

**干什么：** 多个 Excel 文件拼成一张总表。比如 1~6 月报表各一个文件 → 合成上半年总表。

**三步搞定：**
1. 📁 拖拽或选择多个 Excel 文件（至少 2 个）
2. 🚀 点执行
3. 📥 下载合并后的总表

**💡 举个栗子：**
- 文件 A：上半年销售（6 行）
- 文件 B：下半年销售（6 行）
→ 合并后变成 12 行的大表。

**⚠️ 注意：** 按行纵向拼接。如果列名不一样，空位会自动填 NaN。建议先确保两个文件列名一致。
""",
    "filter": """
### 🎯 数据筛选 — 怎么用？

**干什么：** 从大表里筛出你关心的那部分数据。比如「金额大于 1000 的行」「区域包含华东的行」。

**三步搞定：**
1. 📁 上传 Excel 文件
2. 🔽 选一列 → 选条件 → 填值
3. 🚀 点执行 → 下载筛选结果

**💡 可以这样玩：**
| 场景 | 列 | 条件 | 值 |
|------|-----|------|-----|
| 找大额订单 | 金额 | 大于 (>) | 100000 |
| 排除某个区域 | 区域 | 不等于 (!=) | 华北 |
| 模糊搜索 | 产品名 | 包含 | 手机 |
| 找空白单元格 | 备注 | 为空 | (不用填) |

**⚠️ 注意：** 大于/小于只能用在数字列上。
""",
    "match": """
### 🔗 匹配对比 — 怎么用？

**干什么：** 两张表按共同列关联起来，就像 Excel 的 VLOOKUP。比如左表有「员工ID+姓名」，右表有「员工ID+部门」→ 合在一起看哪个员工在哪个部门。

**三步搞定：**
1. 📁 左边上传主表，右边上传对照表
2. 🔗 分别选两边的匹配列（通常是 ID 列）
3. 🔘 选匹配方式：
   - **left** — 保留左表全部行（推荐，跟 VLOOKUP 一样）
   - **inner** — 只保留两边都匹配上的
   - **outer** — 两边都保留，没匹配到的填空
4. 🚀 点执行 → 下载

**💡 举个栗子：**
- 左表（员工表）：E001张三 / E002李四 / E004赵六
- 右表（部门表）：E001技术部 / E003产品部 / E006市场部
- 按员工ID → left 匹配 → 张三=技术部，李四=没匹配上，赵六=没匹配上

**⚠️ 注意：** 如果两边列名不同也没关系，分别选就行。
""",
    "pivot": """
### 📐 透视汇总 — 怎么用？

**干什么：** 跟 Excel 的透视表一个意思——按类别自动汇总。比如「每个产品总销售额」「每个区域平均订单量」。

**三步搞定：**
1. 📁 上传 Excel 文件
2. 🔽 选行标签（按什么分类，比如「产品」）
3. 🔢 选数值列（要汇总的数字，比如「销售额」）
4. 🧮 选汇总方式（求和/计数/平均值/最大/最小）
5. 🚀 点执行 → 下载

**💡 举个栗子：**
原始数据 20 行，包含 4 种产品 × 5 个月 → 选行标签「产品」，数值「销售额」，汇总「求和」→ 输出 4 行，每种产品一个总销售额。

**⚠️ 注意：** 数值列只对数字列生效。如果列全是文字，改用「计数」。
""",
    "summary": """
### 📋 自动汇总 — 怎么用？

**干什么：** 按某一列分组，每组底下自动插入「小计」行，末尾插入「合计」行，带颜色格式。

**三步搞定：**
1. 📁 上传 Excel 文件
2. 🔽 选分组列（比如「部门」）
3. ☑️ 勾选要汇总的数值列（可多选，比如「预算」「实际支出」）
4. 🚀 点执行 → 下载

**💡 输出长这样：**
```
技术部  Q1  500000
技术部  Q2  550000
📌 技术部 小计  1,050,000    ← 蓝色底
市场部  Q1  300000
📌 市场部 小计  300,000      ← 蓝色底
💰 总计 / Grand Total  xxxxx ← 金色底
```

**⚠️ 注意：** 下载后用 Excel 打开才能看到完整的颜色格式。
""",
    "condfmt": """
### 🎨 条件格式 — 怎么用？

**干什么：** 把符合条件的单元格标成彩色。比如「不及格的标红」「超过目标的标绿」。

**三步搞定：**
1. 📁 上传 Excel 文件
2. 🔽 选要标记的列 → 选条件（大于/小于/等于/包含）→ 填入阈值 → 选颜色
3. 🚀 点执行 → 下载

**💡 可以这样玩：**
| 场景 | 列 | 条件 | 阈值 | 颜色 |
|------|-----|------|------|------|
| 找不及格 | 数学 | 小于 | 60 | 红色 |
| 找超标 | 费用 | 大于 | 10000 | 橙色 |
| 含关键词 | 备注 | 包含 | 紧急 | 黄色 |

**⚠️ 注意：** 只标记符合条件的那个单元格，不会整行变色。下载后用 Excel 打开看效果最好。
""",
    "formula": """
### 🧮 添加公式列 — 怎么用？

**干什么：** 给表格最右边加一列 Excel 公式。下载后用 Excel 打开，公式自动算出结果。

**三步搞定：**
1. 📁 上传 Excel 文件
2. ✏️ 输入新列名称（比如「总价」）+ Excel 公式（比如 `=B2*C2`）
3. 🚀 点执行 → 下载 → 用 Excel 打开看结果

**💡 常用公式：**
| 想算什么 | 公式举例 |
|----------|----------|
| 单价×数量 | `=B2*C2` |
| 含税价 | `=B2*1.13` |
| 条件判断 | `=IF(C2>10, "多", "少")` |
| 求和 | `=SUM(B2:B100)` |

**⚠️ 注意：** 公式必须以 `=` 开头。页面预览不显示计算结果，必须下载后 Excel 打开才自动算。
""",
    "attend": """
### 👔 考勤工资 — 怎么用？

**干什么：** 上传任何考勤打卡记录，自动算迟到/早退/缺勤，生成工资表。

**四步搞定：**
1. 📥 先下载模板看格式（也可以用你自己的表，列名不一样没关系）
2. 📁 上传你的考勤数据
3. 🔗 **列映射：** 告诉工具你的表里哪列对应「姓名」「日期」「上班」「下班」「基本工资」
4. 🏢 设置公司规则：几点上班/下班 + 迟到罚款 + 缺勤扣款
5. 🚀 点执行 → 下载工资表

**💡 支持的公司系统：**
| 系统 | 常见列名 |
|------|---------|
| 钉钉 | 姓名、考勤日期、签到时间、签退时间 |
| 企业微信 | 姓名、日期、上班、下班 |
| 飞书 | 员工、日期、签到、签退 |
| 其他 | 不管叫什么，映射一下就能用 ✅ |

**📊 输出工资表包含：**
```
姓名 | 出勤天数 | 迟到次数 | 早退次数 | 缺勤天数 | 基本工资 | 迟到扣款 | 缺勤扣款 | 实发工资
```
扣款单元格自动标红，工资列自动标绿。

**⚠️ 注意：**
- 打卡时间支持 08:55、08:55:30、2026-08-01 08:55:00 等格式
- 全天上班+下班都没打卡 → 算缺勤
- 只有下班没打或只有上班没打 → 不算缺勤（可能是忘了打卡）
""",
}

def show_guide(func_id: str):
    """显示当前功能的使用指南。上传文件前默认展开，上传后折叠。"""
    guide = GUIDES.get(func_id, "")
    if guide:
        with st.expander("📖 使用指南 / 怎么用？", expanded=True):
            st.markdown(guide)

# ===========================================================================
# 动态参数区
# ===========================================================================
if cur_func:
    st.markdown('<p class="step-label">⚙️ 第四步：配置参数</p>', unsafe_allow_html=True)
    show_guide(cur_func)

params_ok = False

# ---- 去重 ----
if cur_func == "dedup":
    if df is not None:
        cols_list = df.columns.tolist()
        sel_cols = st.multiselect("选择去重依据列：", cols_list, default=cols_list[:1] if cols_list else [])
        if sel_cols:
            st.info(f"将对 {len(sel_cols)} 列进行去重，保留首行")
            params_ok = True
        st.caption(f"📋 数据：{df.shape[0]} 行 × {df.shape[1]} 列")
    else:
        st.info("👆 请先上传文件")

# ---- 拆分 ----
elif cur_func == "split":
    if df is not None:
        c1, c2 = st.columns(2)
        with c1:
            split_col = st.selectbox("拆分依据列：", df.columns.tolist())
        with c2:
            split_mode = st.radio("模式：", ["单文件多Sheet", "多个文件打包ZIP"], horizontal=True)
        if split_col:
            st.info(f"该列有 {df[split_col].nunique()} 个不同值 → 拆成 {df[split_col].nunique()} 份")
            params_ok = True
        st.caption(f"📋 数据：{df.shape[0]} 行 × {df.shape[1]} 列")
    else:
        st.info("👆 请先上传文件")

# ---- 合并 ----
elif cur_func == "merge":
    if uploaded_files:
        st.write(f"已上传 **{len(uploaded_files)}** 个文件：")
        for i, f in enumerate(uploaded_files, 1):
            st.caption(f"  {i}. {f.name}")
        params_ok = True
    else:
        st.info("👆 请上传至少 2 个文件")

# ---- 筛选 ----
elif cur_func == "filter":
    if df is not None:
        c1, c2, c3 = st.columns(3)
        with c1:
            flt_col = st.selectbox("筛选列：", df.columns.tolist())
        with c2:
            flt_op = st.selectbox("条件：", [
                "等于 (=)", "不等于 (!=)", "大于 (>)", "小于 (<)",
                "大于等于 (>=)", "小于等于 (<=)", "包含", "不包含", "为空", "不为空",
            ])
        with c3:
            if flt_op not in ("为空", "不为空"):
                flt_val = st.text_input("值：", placeholder="输入比较值")
            else:
                flt_val = ""
                params_ok = True
        if flt_val or flt_op in ("为空", "不为空"):
            st.info(f"将筛选：{flt_col} {flt_op} {flt_val}")
            params_ok = True
        st.caption(f"📋 数据：{df.shape[0]} 行 × {df.shape[1]} 列")
    else:
        st.info("👆 请先上传文件")

# ---- 匹配对比 ----
elif cur_func == "match":
    if df is not None and df2 is not None:
        c1, c2, c3 = st.columns(3)
        with c1:
            m_col1 = st.selectbox("主表匹配列：", df.columns.tolist(), key="mcol1")
        with c2:
            m_col2 = st.selectbox("对照表匹配列：", df2.columns.tolist(), key="mcol2")
        with c3:
            m_how = st.selectbox("匹配方式：", ["left", "inner", "right", "outer"],
                                 help="left=保留左表全部 | inner=只保留匹配上的 | outer=都保留")
        st.info(f"将按「{m_col1}」↔「{m_col2}」进行 {m_how} 匹配")
        st.caption(f"📋 主表：{df.shape[0]}行×{df.shape[1]}列 | 对照表：{df2.shape[0]}行×{df2.shape[1]}列")
        params_ok = True
    else:
        st.info("👆 请在右侧分别上传主表和对照表")

# ---- 透视汇总 ----
elif cur_func == "pivot":
    if df is not None:
        # 找出数值列
        num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        all_cols = df.columns.tolist()
        c1, c2, c3 = st.columns(3)
        with c1:
            pv_idx = st.selectbox("行标签（分组列）：", all_cols)
        with c2:
            pv_val = st.selectbox("数值列：", num_cols if num_cols else all_cols)
        with c3:
            pv_agg = st.selectbox("汇总方式：", ["求和", "计数", "平均值", "最大值", "最小值"])
        params_ok = True
        st.caption(f"📋 数据：{df.shape[0]} 行 × {df.shape[1]} 列")
    else:
        st.info("👆 请先上传文件")

# ---- 自动汇总 ----
elif cur_func == "summary":
    if df is not None:
        all_cols = df.columns.tolist()
        num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        if not num_cols:
            st.warning("表格中没有数值列，无法汇总")
        else:
            c1, c2 = st.columns(2)
            with c1:
                sm_grp = st.selectbox("分组列：", all_cols)
            with c2:
                sm_vals = st.multiselect("汇总列（可多选）：", num_cols, default=num_cols[:1])
            if sm_grp and sm_vals:
                st.info(f"按「{sm_grp}」分组，汇总 {len(sm_vals)} 列 → 每组生成小计行 + 末尾合计")
                params_ok = True
        st.caption(f"📋 数据：{df.shape[0]} 行 × {df.shape[1]} 列")
    else:
        st.info("👆 请先上传文件")

# ---- 条件格式 ----
elif cur_func == "condfmt":
    if df is not None:
        all_cols = df.columns.tolist()
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            cf_col = st.selectbox("要标记的列：", all_cols)
        with c2:
            cf_cond = st.selectbox("条件：", ["大于", "小于", "等于", "大于等于", "小于等于", "包含"])
        with c3:
            cf_val = st.text_input("阈值：", placeholder="输入比较值")
        with c4:
            cf_color = st.selectbox("标记颜色：", ["红色", "绿色", "黄色", "蓝色", "橙色", "紫色"])
        if cf_val:
            st.info(f"{cf_col} 列中 {cf_cond}「{cf_val}」的单元格将标记为 {cf_color}")
            params_ok = True
        st.caption(f"📋 数据：{df.shape[0]} 行 × {df.shape[1]} 列")
    else:
        st.info("👆 请先上传文件")

# ---- 公式列 ----
elif cur_func == "formula":
    if df is not None:
        c1, c2 = st.columns(2)
        with c1:
            fm_name = st.text_input("新列名称：", placeholder="例如：总价")
        with c2:
            fm_formula = st.text_input("Excel 公式（= 开头）：", placeholder="例如：=B2*C2")
        if fm_name and fm_formula:
            if not fm_formula.startswith("="):
                st.warning("公式需以 = 开头")
            else:
                params_ok = True
        st.caption(f"📋 数据：{df.shape[0]} 行 × {df.shape[1]} 列")
    else:
        st.info("👆 请先上传文件")

# ---- 考勤工资 ----
elif cur_func == "attend":
    if df is not None:
        all_cols = df.columns.tolist()

        # ---- 列映射：自己选你的表里哪列对应哪个字段 ----
        st.markdown("**🔗 列映射：** 你的表里哪列对应哪个字段？")
        col_map = {}
        map_cols = st.columns(5)
        field_info = [
            ("姓名", "员工姓名/工号列"),
            ("日期", "考勤日期列"),
            ("上班打卡", "上班签到时间列"),
            ("下班打卡", "下班签退时间列"),
            ("基本工资", "底薪列（没有可不选）"),
        ]
        for i, (field, help_text) in enumerate(field_info):
            with map_cols[i]:
                choices = all_cols.copy()
                if field == "基本工资":
                    choices.insert(0, "（无此列，用默认工资）")
                col_map[field] = st.selectbox(
                    f"{field}：", choices,
                    index=all_cols.index(field) if field in all_cols else (
                        0 if field == "基本工资" else None
                    ),
                    help=help_text,
                    key=f"map_{field}",
                )

        # 验证映射
        required_fields = ["姓名", "日期", "上班打卡", "下班打卡"]
        mapped_ok = all(
            col_map[f] in all_cols for f in required_fields
        )
        salary_has_col = col_map["基本工资"] in all_cols

        if not mapped_ok:
            st.warning("⚠️ 请把前四项映射到表中存在的列")
        else:
            attr_map = {}  # collect for display
            for f in required_fields:
                attr_map[f] = col_map[f]
            st.success(
                f"✅ 映射成功：" +
                " | ".join(f"{k} →「{v}」" for k, v in attr_map.items()) +
                (f" | 基本工资 →「{col_map['基本工资']}」" if salary_has_col else " | 基本工资 → 使用默认值")
            )

            # ---- 公司规则 ----
            st.markdown("**🏢 公司考勤规则：**")
            c1, c2, c3, c4, c5 = st.columns(5)
            with c1:
                at_start = st.text_input("上班时间：", value="09:00")
            with c2:
                at_end = st.text_input("下班时间：", value="18:00")
            with c3:
                at_late = st.number_input("迟到扣款/次：", value=50.0, step=10.0)
            with c4:
                at_absent = st.number_input("缺勤扣款/天：", value=300.0, step=50.0)
            with c5:
                at_base_default = st.number_input(
                    "默认基本工资：", value=5000.0, step=500.0,
                    help="只有表中没有工资列时才用这个值",
                )

            st.info(
                "计算规则：上班打卡晚于截止时间=迟到 | "
                "下班打卡早于截止时间=早退 | "
                "全天没打卡=缺勤 | "
                "实发 = 基本工资 − 迟到扣款 − 缺勤扣款"
            )
            params_ok = True

        st.caption(f"📋 数据预览：{df.shape[0]} 行 × {df.shape[1]} 列")
        with st.expander("🔍 查看数据前 10 行"):
            st.dataframe(df.head(10), use_container_width=True)
    else:
        st.info("👆 请先上传文件（或先下载模板填写后上传）")

st.divider()

# ===========================================================================
# 执行按钮
# ===========================================================================
st.markdown('<p class="step-label">🚀 第五步：执行处理</p>', unsafe_allow_html=True)

exec_col, _ = st.columns([1, 3])
with exec_col:
    exec_clicked = st.button(
        "⚡ 开始处理", disabled=not (cur_func and params_ok),
        type="primary", use_container_width=True,
    )

# ===========================================================================
# 处理逻辑
# ===========================================================================
if exec_clicked and cur_func and params_ok:
    result_bytes = None
    result_ext = ".xlsx"
    success_msg = ""
    func_cn = {
        "dedup": "去重", "split": "拆分", "merge": "合并", "filter": "筛选",
        "match": "匹配", "pivot": "透视", "summary": "汇总", "condfmt": "条件格式",
        "formula": "公式列", "attend": "考勤工资",
    }
    prefix = func_cn.get(cur_func, "结果")

    with st.spinner("⏳ 处理中..."):
        try:
            # --- 去重 ---
            if cur_func == "dedup":
                rdf, removed = do_dedup(df, sel_cols)
                result_bytes = df_to_excel(rdf)
                success_msg = f"✅ 去掉了 {removed} 行重复，剩余 {len(rdf)} 行"

            # --- 拆分 ---
            elif cur_func == "split":
                result_bytes = do_split(df, split_col, split_mode)
                result_ext = ".zip"
                success_msg = f"✅ 已拆分为 {df[split_col].nunique()} 份"

            # --- 合并 ---
            elif cur_func == "merge":
                rdf = do_merge(uploaded_files)
                if rdf is not None:
                    result_bytes = df_to_excel(rdf)
                    success_msg = f"✅ 合并完成，共 {len(rdf)} 行 × {rdf.shape[1]} 列"
                else:
                    st.error("所有文件读取失败或为空")

            # --- 筛选 ---
            elif cur_func == "filter":
                rdf = do_filter(df, flt_col, flt_op, flt_val)
                result_bytes = df_to_excel(rdf)
                success_msg = f"✅ 筛选完成：{len(rdf)} 行满足条件（从 {len(df)} 行中）"

            # --- 匹配 ---
            elif cur_func == "match":
                rdf = do_match(df, df2, m_col1, m_col2, m_how)
                result_bytes = df_to_excel(rdf)
                success_msg = f"✅ 匹配完成：{len(rdf)} 行 × {rdf.shape[1]} 列"

            # --- 透视 ---
            elif cur_func == "pivot":
                rdf = do_pivot(df, pv_idx, pv_val, pv_agg)
                result_bytes = df_to_excel(rdf)
                success_msg = f"✅ 透视完成：{len(rdf)} 行汇总"

            # --- 汇总 ---
            elif cur_func == "summary":
                result_bytes = do_auto_summary(df, sm_grp, sm_vals)
                success_msg = f"✅ 汇总完成：按「{sm_grp}」分组，含小计+合计"

            # --- 条件格式 ---
            elif cur_func == "condfmt":
                result_bytes = do_conditional_format(df, cf_col, cf_cond, cf_val, cf_color)
                success_msg = f"✅ 已标记：{cf_col} 列中 {cf_cond}「{cf_val}」的单元格为 {cf_color}"

            # --- 公式列 ---
            elif cur_func == "formula":
                result_bytes = do_formula(df, fm_name, fm_formula)
                success_msg = f"✅ 已添加公式列「{fm_name}」"

            # --- 考勤 ---
            elif cur_func == "attend":
                result_bytes = do_attendance_salary(
                    df,
                    work_start=at_start,
                    work_end=at_end,
                    late_penalty=at_late,
                    absent_penalty=at_absent,
                    default_salary=at_base_default,
                    name_col=col_map["姓名"],
                    date_col=col_map["日期"],
                    start_col=col_map["上班打卡"],
                    end_col=col_map["下班打卡"],
                    salary_col=col_map["基本工资"] if salary_has_col else None,
                )
                success_msg = "✅ 考勤工资核算完成"

        except Exception as e:
            st.error(f"❌ 处理出错：{e}")
            import traceback
            with st.expander("🔍 错误详情"):
                st.code(traceback.format_exc())

    # 保存结果
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    st.session_state.result_bytes = result_bytes
    st.session_state.result_ext = result_ext
    st.session_state.result_filename = f"{prefix}_{timestamp}{result_ext}"
    st.session_state.success_msg = success_msg

    if success_msg:
        st.success(success_msg)
    st.rerun()

# ===========================================================================
# 下载区
# ===========================================================================
if st.session_state.result_bytes:
    st.divider()
    st.markdown('<p class="step-label">📥 下载结果</p>', unsafe_allow_html=True)
    st.download_button(
        label=f"📥 下载：{st.session_state.result_filename}",
        data=st.session_state.result_bytes,
        file_name=st.session_state.result_filename,
        mime="application/octet-stream",
        use_container_width=True,
    )

# ===========================================================================
# 空状态说明
# ===========================================================================
if not cur_func:
    st.markdown("---")
    st.markdown("""
    ### 💡 使用说明

    | 分类 | 功能 | 一句话 |
    |---|---|---|
    | 📊 数据处理 | 🔍 去重 | 按列删重复行 |
    | | ✂️ 拆分 | 按列拆成多个表 |
    | | 🔗 合并 | 多文件拼成总表 |
    | | 🎯 筛选 | 按条件过滤数据 |
    | | 🔗 匹配 | 两张表 VLOOKUP |
    | 📈 分析汇总 | 📐 透视 | 分类求和/平均 |
    | | 📋 汇总 | 分组小计+合计 |
    | | 🎨 条件格式 | 高亮标记单元格 |
    | 📋 办公模板 | 🧮 公式列 | 添加 Excel 公式 |
    | | 👔 考勤工资 | 打卡→迟到→工资 |

    👆 请在上方选择一个分类和功能开始使用
    """)

st.divider()
st.caption("🔒 所有处理均在浏览器内存中完成，不会保存您的任何数据到服务器。")
st.caption(f"Excel 百宝箱 · 10 项功能 · Powered by Streamlit · {datetime.now().year}")
